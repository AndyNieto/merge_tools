import os
import openpyxl
from datetime import date, datetime

# Define the directory path
directory = os.path.expanduser("~/Documents/Data_Collection")

# Create a subdirectory based on the current date
current_date = date.today().strftime("%Y-%m-%d")
subdirectory_input = os.path.join(directory, "input", "samsung")
subdirectory_output = os.path.join(directory, current_date, "output", "samsung")
os.makedirs(subdirectory_output, exist_ok=True)

# Get the list of Excel files to merge from the input subdirectory
excel_files = [file for file in os.listdir(subdirectory_input) if file.endswith('.xlsx')]

# Create a new workbook to store the merged data
merged_workbook = openpyxl.Workbook()

# Iterate over the Excel files in the input subdirectory
for file in excel_files:
    # Open the workbook
    workbook = openpyxl.load_workbook(os.path.join(subdirectory_input, file))
    
    # Iterate over the sheets in the workbook
    for sheet_name in workbook.sheetnames:
        # Get the sheet
        sheet = workbook[sheet_name]
        
        # Create a new sheet in the merged workbook if it doesn't exist
        if sheet_name not in merged_workbook.sheetnames:
            merged_sheet = merged_workbook.create_sheet(title=sheet_name)
        else:
            merged_sheet = merged_workbook[sheet_name]
        
        # Copy the data from the source sheet to the merged sheet
        for row in sheet.iter_rows(min_row=3, values_only=True):
            # Exclude the first column (column A)
            merged_sheet.append(row[1:])

# Remove the default sheet created by openpyxl
del merged_workbook['Sheet']

# Save the merged workbook in the output subdirectory
output_file_name = "merged_file_" + datetime.now().strftime("%Y-%m-%d_T%H-%M-%S") + ".xlsx"
merged_file_path = os.path.join(subdirectory_output, output_file_name)

# Save the merged workbook in the output subdirectory
merged_workbook.save(merged_file_path)
